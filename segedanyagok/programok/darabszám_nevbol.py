import pandas as pd
import openpyxl
import re
import os

r'''
input_path = r"D:\Nagy_Gabor\SZVG_cuccok\csavarok\alapfajlok\din125.xls"
excel_szurt_path = r"D:\Nagy_Gabor\SZVG_cuccok\csavarok\szurtek\szurt_csavarok_125.xlsx"
excel_maradek_path = r"D:\Nagy_Gabor\SZVG_cuccok\csavarok\maradek\maradek_csavarok_125.xlsx"


tiltott_mintak = ["100/csomag", "100db/cs", "100db/csomag", "100db/csom"]

wb = openpyxl.load_workbook(input_path)
sheet = wb.active  

wb_szurt = openpyxl.Workbook()
sheet_szurt = wb_szurt.active
wb_maradek = openpyxl.Workbook()
sheet_maradek = wb_maradek.active

header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
sheet_szurt.append(header)
sheet_maradek.append(header)

for row in sheet.iter_rows(min_row=2, values_only=True):
    csavar_nev = str(row[0]) if row[0] else ""
    csavar_nev_lower = csavar_nev.lower()
    csavar_nev = csavar_nev.replace("Hlf.rm.csavar", "Hatlapfejű részmenetes csavar")

    if any(minta in csavar_nev_lower for minta in tiltott_mintak):
        for minta in tiltott_mintak:
            csavar_nev = re.sub(minta, " ", csavar_nev, flags=re.IGNORECASE)

        uj_sor = list(row)
        uj_sor[0] = csavar_nev.strip()
        sheet_szurt.append(uj_sor)
    else:
        uj_sor = list(row)
        uj_sor[0] = csavar_nev.strip()
        sheet_maradek.append(uj_sor)


wb_szurt.save(excel_szurt_path)
wb_maradek.save(excel_maradek_path)

print(f"Kész!\nSzűrt Excel: {excel_szurt_path}\nMaradék Excel: {excel_maradek_path}")
'''

input_path = r"D:\Nagy_Gabor\SZVG_cuccok\csavarok\alapfajlok\din9021.xls"
excel_szurt_path = r"D:\Nagy_Gabor\SZVG_cuccok\csavarok\szurtek\szurt_csavarok_9021.xlsx"
excel_maradek_path = r"D:\Nagy_Gabor\SZVG_cuccok\csavarok\maradek\maradek_csavarok_9021.xlsx"

tiltott_mintak = ["100/csomag", "100db/cs", "100db/csomag", "100db/csom"]

df = pd.read_excel(input_path, engine="xlrd", header=0)

df_szurt = pd.DataFrame(columns=df.columns)
df_maradek = pd.DataFrame(columns=df.columns)

for idx, row in df.iloc[1:].iterrows():  
    csavar_nev = str(row.iloc[1]) if pd.notna(row.iloc[1]) else "" 
    csavar_nev_lower = csavar_nev.lower()
    #csavar_nev = csavar_nev.replace("Sf.bkny.csavar", "Süllyesztettfejű belsőkulcsnyílású csavar")
    if any(minta in csavar_nev_lower for minta in tiltott_mintak):
        for minta in tiltott_mintak:
            csavar_nev = re.sub(minta, " ", csavar_nev, flags=re.IGNORECASE)

        row.iloc[1] = csavar_nev.strip()
        df_szurt.loc[len(df_szurt)] = row
    else:
        row.iloc[1] = csavar_nev.strip()
        df_maradek.loc[len(df_maradek)] = row

os.makedirs(os.path.dirname(excel_szurt_path), exist_ok=True)
df_szurt.to_excel(excel_szurt_path, index=False)
df_maradek.to_excel(excel_maradek_path, index=False)

print(f"pipa")