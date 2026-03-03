from openpyxl import load_workbook

input_fajl = "/Users/nagygabor/Library/Mobile Documents/com~apple~CloudDocs/munka_fajlok/majdnemkesz_csavarok/kesz_7991.xlsx"
output_fajl = "/Users/nagygabor/Library/Mobile Documents/com~apple~CloudDocs/munka_fajlok/kesz_csav/DIN_7991.xlsx"
oszlop = 'B'
kezdo_sor = 2
prefix = "DIN 7991 "
sima ="https://csavar.bolt.hu/images/virtuemart/product/belso_kulcsnyilasu_sullyesztettfeju.jpg"
#kepurl_a2 = "https://wimg.b-cdn.net/8f48a7589834c62d468712353a5cd4a2/categories/din933a4-dd0ad.jpg?width=280&height=280"
#kepurl_a4 = "https://wimg.b-cdn.net/8f48a7589834c62d468712353a5cd4a2/categories/din933a1-040d3.jpg?width=280&height=280"
kepurl_bl = "https://wimg.b-cdn.net/8f48a7589834c62d468712353a5cd4a2/products/din7991-e8212.jpg?width=144&height=144"
#kep_rezes = "https://csavarda.hu/storage/product_images/934sarga.png"

wb = load_workbook(input_fajl)
ws = wb.active

for sor in range(kezdo_sor, ws.max_row + 1):
    cella = ws[f"{oszlop}{sor}"]
    if cella.value:
        uj_nev = prefix + str(cella.value)
        cella.value = uj_nev

        if "BL" in uj_nev.upper():
            ws[f"T{sor}"].value = kepurl_bl
        else:
            ws[f"T{sor}"].value = sima

        
wb.save(output_fajl)

print(f"pipa")
